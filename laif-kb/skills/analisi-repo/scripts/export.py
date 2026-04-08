#!/usr/bin/env python3
"""
Export dati analisi in SQLite, Excel e Markdown.

Uso:
    python3 export.py --input output/ [--db output/analysis.db] [--excel output/analysis.xlsx]
"""

import argparse
import json
import sqlite3
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_analyses(input_dir: str) -> list:
    """Carica tutti i JSON di analisi dalla directory."""
    analyses = []
    for f in sorted(Path(input_dir).glob("*.json")):
        if f.name == "cross_analysis.json":
            continue
        try:
            analyses.append(json.loads(f.read_text()))
        except json.JSONDecodeError:
            pass
    return analyses


# --- SQLite ---

def create_sqlite(analyses: list, db_path: str):
    """Crea database SQLite con tutte le tabelle."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # Tabella repos
    c.execute("""CREATE TABLE IF NOT EXISTS repos (
        name TEXT PRIMARY KEY,
        template_version TEXT,
        baseline_version TEXT,
        delta_minor INTEGER,
        last_commit_date TEXT,
        last_commit_author TEXT,
        custom_pages INTEGER,
        custom_modals INTEGER,
        custom_tables INTEGER,
        custom_endpoints INTEGER,
        ds_components_used INTEGER,
        custom_components INTEGER,
        extra_deps_backend INTEGER,
        extra_deps_frontend INTEGER
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS custom_tables (
        repo TEXT, table_name TEXT, class_name TEXT, file TEXT,
        PRIMARY KEY (repo, table_name)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS custom_endpoints (
        repo TEXT, method TEXT, path TEXT, type TEXT, file TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS custom_pages (
        repo TEXT, route TEXT, depth INTEGER,
        PRIMARY KEY (repo, route)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS ds_components_usage (
        repo TEXT, component TEXT, count INTEGER,
        PRIMARY KEY (repo, component)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS custom_components (
        repo TEXT, file TEXT,
        PRIMARY KEY (repo, file)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS extra_deps_backend (
        repo TEXT, package TEXT, version TEXT,
        PRIMARY KEY (repo, package)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS extra_deps_frontend (
        repo TEXT, package TEXT, version TEXT,
        PRIMARY KEY (repo, package)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS contributors (
        repo TEXT, name TEXT, commits INTEGER,
        PRIMARY KEY (repo, name)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS background_tasks (
        repo TEXT, type TEXT, file TEXT
    )""")

    # Popola
    for a in analyses:
        repo = a["repo"]
        v = a.get("versioning", {})
        dm = a.get("datamodel", {})
        fp = a.get("frontend_pages", {})
        fc = a.get("frontend_components", {})
        api = a.get("api_routes", {})
        deps = a.get("dependencies", {})
        gs = a.get("git_stats", {})
        delta = v.get("delta", {})
        last = gs.get("last_commit", {})

        c.execute("INSERT OR REPLACE INTO repos VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            repo, v.get("template_version"), v.get("baseline_version"),
            delta.get("minor") if delta else None,
            last.get("date", "")[:10] if last else None,
            last.get("author") if last else None,
            fp.get("custom_page_count", 0), fp.get("custom_modal_count", 0),
            dm.get("custom_table_count", 0), api.get("custom_endpoint_count", 0),
            fc.get("ds_component_count", 0), fc.get("custom_component_count", 0),
            deps.get("extra_backend_count", 0), deps.get("extra_frontend_count", 0),
        ))

        for t in dm.get("custom_tables", []):
            c.execute("INSERT OR REPLACE INTO custom_tables VALUES (?,?,?,?)",
                      (repo, t["table"], t.get("class", ""), t.get("file", "")))

        for ep in api.get("custom_endpoints", []):
            c.execute("INSERT INTO custom_endpoints VALUES (?,?,?,?,?)",
                      (repo, ep["method"], ep["path"], ep["type"], ep["file"]))

        for p in fp.get("custom_pages", []):
            c.execute("INSERT OR REPLACE INTO custom_pages VALUES (?,?,?)",
                      (repo, p["route"], p["depth"]))

        for comp, count in fc.get("ds_components_used", {}).items():
            c.execute("INSERT OR REPLACE INTO ds_components_usage VALUES (?,?,?)",
                      (repo, comp, count))

        for comp_file in fc.get("custom_components", []):
            c.execute("INSERT OR REPLACE INTO custom_components VALUES (?,?)",
                      (repo, comp_file))

        for dep in deps.get("extra_backend", []):
            c.execute("INSERT OR REPLACE INTO extra_deps_backend VALUES (?,?,?)",
                      (repo, dep["name"], dep["version"]))

        for dep in deps.get("extra_frontend", []):
            c.execute("INSERT OR REPLACE INTO extra_deps_frontend VALUES (?,?,?)",
                      (repo, dep["name"], dep["version"]))

        for cont in gs.get("contributors", []):
            c.execute("INSERT OR REPLACE INTO contributors VALUES (?,?,?)",
                      (repo, cont["name"], cont["commits"]))

        for task in a.get("background_tasks", {}).get("background_tasks", []):
            c.execute("INSERT INTO background_tasks VALUES (?,?,?)",
                      (repo, task["type"], task["file"]))

    conn.commit()
    conn.close()
    print(f"SQLite: {db_path} ({len(analyses)} repos)")


# --- Excel ---

def create_excel(analyses: list, excel_path: str):
    """Genera Excel multi-foglio."""
    if not HAS_OPENPYXL:
        print("SKIP Excel: installa openpyxl (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()

    # Foglio repos
    ws = wb.active
    ws.title = "Repos"
    headers = ["Repo", "Template Ver.", "Delta Minor", "Pagine", "Modali",
               "Tabelle", "Endpoint", "DS Components", "Custom Components",
               "Extra BE", "Extra FE", "Ultimo Commit"]
    ws.append(headers)
    for a in analyses:
        v = a.get("versioning", {})
        dm = a.get("datamodel", {})
        fp = a.get("frontend_pages", {})
        fc = a.get("frontend_components", {})
        api = a.get("api_routes", {})
        deps = a.get("dependencies", {})
        gs = a.get("git_stats", {})
        delta = v.get("delta", {})
        last = gs.get("last_commit", {})
        ws.append([
            a["repo"], v.get("template_version"),
            delta.get("minor") if delta else None,
            fp.get("custom_page_count", 0), fp.get("custom_modal_count", 0),
            dm.get("custom_table_count", 0), api.get("custom_endpoint_count", 0),
            fc.get("ds_component_count", 0), fc.get("custom_component_count", 0),
            deps.get("extra_backend_count", 0), deps.get("extra_frontend_count", 0),
            last.get("date", "")[:10] if last else "",
        ])

    # Foglio DS Components
    ws2 = wb.create_sheet("DS Components")
    ws2.append(["Repo", "Componente", "Utilizzi"])
    for a in analyses:
        for comp, count in a.get("frontend_components", {}).get("ds_components_used", {}).items():
            ws2.append([a["repo"], comp, count])

    # Foglio Dipendenze Extra
    ws3 = wb.create_sheet("Extra Deps")
    ws3.append(["Repo", "Tipo", "Pacchetto", "Versione"])
    for a in analyses:
        for dep in a.get("dependencies", {}).get("extra_backend", []):
            ws3.append([a["repo"], "backend", dep["name"], dep["version"]])
        for dep in a.get("dependencies", {}).get("extra_frontend", []):
            ws3.append([a["repo"], "frontend", dep["name"], dep["version"]])

    # Foglio Custom Tables
    ws4 = wb.create_sheet("Custom Tables")
    ws4.append(["Repo", "Tabella", "Classe", "File"])
    for a in analyses:
        for t in a.get("datamodel", {}).get("custom_tables", []):
            ws4.append([a["repo"], t["table"], t.get("class", ""), t.get("file", "")])

    wb.save(excel_path)
    print(f"Excel: {excel_path} ({len(analyses)} repos)")


def main():
    parser = argparse.ArgumentParser(description="Export analisi LAIF")
    parser.add_argument("--input", required=True, help="Directory con i JSON di analisi")
    parser.add_argument("--db", default=None, help="Path output SQLite")
    parser.add_argument("--excel", default=None, help="Path output Excel")
    args = parser.parse_args()

    input_dir = args.input
    analyses = load_analyses(input_dir)
    if not analyses:
        print("Nessuna analisi trovata")
        return

    db_path = args.db or str(Path(input_dir) / "analysis.db")
    excel_path = args.excel or str(Path(input_dir) / "analysis.xlsx")

    create_sqlite(analyses, db_path)
    create_excel(analyses, excel_path)


if __name__ == "__main__":
    main()
